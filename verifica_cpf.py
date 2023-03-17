import pyautogui
from time import sleep
from pywinauto.keyboard import send_keys
from openpyxl import load_workbook
from tkinter import messagebox
from selenium import webdriver
import time
import pyperclip

linha = 1
j = 0



#verifica se a planilha esta correta
try:
    workbook = load_workbook(r"C:\Users\natalia.policeno\Desktop\exel_portifolio\portifolio.xlsx")
except: 
    messagebox.showinfo(title="ERRO",message="O caminho da planilha está errado!")  
    exit()
    
worksheet = workbook.active

#por conta do site gerador de cpf ser um site protejido,nao funciona com selenium.Essa é uma das vantagens da automaçao pyautugui,em contrapartida é mais demorado.
def entrar():

    pyautogui.press('win')
    sleep(2)

    pyautogui.press('g')
    sleep(2)

    pyautogui.press('enter')
    sleep(2)
    
    pyautogui.write('gerador de cpf')
    sleep(3)
    
    pyautogui.press('enter')
    sleep(2)
    gerar_cpf()
    
def gerar_cpf():
    sleep(2)
    #clica no site "4devs.com.br"
    pyautogui.click(185,311) 
    sleep(2)
    
    #clica em gerar cpf
    pyautogui.click(766,402)
    sleep(2)
    
    #clica no campo do cpf
    pyautogui.click(576,496)
    sleep(2)
    
    #copia as informaçoes
    pyautogui.hotkey('ctrl', 'c')
    sleep(2)
    
    cpf = pyperclip.paste()
    
    #manda para a planilha
    worksheet.cell(row=linha + 1, column=1).value = cpf
    
    verificacao(cpf)
    
def verificacao(cpf):
    len(cpf)

    cpf = cpf.replace('.','').replace('-','')

    arr_num = []
    for num in cpf:
        arr_num.append(int(num))
        
    print(cpf)
    print(arr_num)
        
    primeira_soma = 0
    for i in range(0,9):
        primeira_soma += arr_num[i] * (10 - i)
        
    primeiro_digito_verificador = (primeira_soma * 10) % 11
        
    if primeiro_digito_verificador == 10:
            primeiro_digito_verificador = 0

    segunda_soma = 0
    for i in range(0,10):
        segunda_soma += arr_num[i] * (11 - i)
        
    segunda_soma += primeiro_digito_verificador * 2
    
    segundo_digito_verificador = (segunda_soma * 10) % 11

    if segundo_digito_verificador == 10:
        segundo_digito_verificador = 0
        
    if (primeiro_digito_verificador == arr_num[9] and segundo_digito_verificador == arr_num[10]):
        worksheet.cell(row=linha + 1, column=2).value ='valido'

    else:
        worksheet.cell(row=linha + 1, column=2).value ='invalido'
        
     
def iniciar():
    global linha  
    
    #se nao tiver informaçoes vai continuar o loop
    for j in range(0,10):
        
        #caso a linha nao tenha sido verificada vai continuar o loop a partir da linha
        if worksheet.cell(row=linha + 1, column=2).value == None:

            gerar_cpf()

            linha = linha + 1
            
            #salva as informaçoes
            workbook.save(r"C:\Users\natalia.policeno\Desktop\exel_portifolio\portifolio.xlsx")

        else:
            linha = linha + 1   

entrar()
iniciar()

messagebox.showinfo(title="Robo", message="Fim")
